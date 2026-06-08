"""
Rebar Shop Drawing QA/QC Auditor  —  v2.0  (Clean Rewrite)
===========================================================
Supports: Ollama (offline) + Claude API (online)
BBS Excel parsed directly — no AI timeout for Excel files
4 focused audit passes instead of one giant prompt

SETUP (Ollama):
  1.  ollama serve
  2.  ollama pull llava:34b
  3.  ollama pull llama3.3:70b   (or llama3.1:8b for speed)
  4.  python app_v2.py
  5.  Open http://localhost:5000

SETUP (Claude API):
  1.  Set CLAUDE_API_KEY below (or env var ANTHROPIC_API_KEY)
  2.  python app_v2.py
"""

import os, sys, json, base64, io, uuid, threading, subprocess, time
from pathlib import Path
from datetime import datetime

# ── Auto-install
REQUIRED = {"flask":"flask","Pillow":"PIL","PyMuPDF":"fitz",
            "requests":"requests","openpyxl":"openpyxl","reportlab":"reportlab","anthropic":"anthropic","openai":"openai","google-generativeai":"google.generativeai"}
def _install():
    for pkg, imp in REQUIRED.items():
        try: __import__(imp)
        except ImportError:
            subprocess.check_call([sys.executable,"-m","pip","install",pkg,
                                   "--quiet","--break-system-packages"],
                                  stderr=subprocess.DEVNULL)
_install()

from flask import Flask, request, jsonify, send_file, Response, stream_with_context
from PIL import Image, ImageDraw, ImageFont
import fitz, requests as rq

# ═══════════════════════════════════════════════════════════
# SECTION 1 — CONFIGURATION
# ═══════════════════════════════════════════════════════════

OLLAMA_URL      = "http://localhost:11434"
# ── AI Engine keys & selection (all updatable at runtime via /set-engine)
_CLAUDE_KEY: str = os.environ.get("ANTHROPIC_API_KEY","")
_OPENAI_KEY: str = os.environ.get("OPENAI_API_KEY","")
_GEMINI_KEY: str = os.environ.get("GOOGLE_API_KEY","")

# Auto-select engine: if API key env var is set, use that engine; else try Ollama
if   _CLAUDE_KEY: _ENGINE: str = "claude"
elif _OPENAI_KEY: _ENGINE: str = "openai"
elif _GEMINI_KEY: _ENGINE: str = "gemini"
else:             _ENGINE: str = "ollama"
MAX_BAR_LEN_MM  = 12000

# Handle PyInstaller bundle path
if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).parent
else:
    BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
REPORT_DIR = BASE_DIR / "reports"
for _d in (UPLOAD_DIR, REPORT_DIR): _d.mkdir(exist_ok=True)

app  = Flask(__name__)
JOBS: dict = {}
LOCK = threading.Lock()

# ── Frontend lookup
FRONTEND = None
for _f in ["rebar_auditor_v2.html","rebar_auditor_frontend_v9.html","rebar_auditor_v2.html","index.html"]:
    _p = BASE_DIR / _f
    if _p.exists(): FRONTEND = _p; break

# ═══════════════════════════════════════════════════════════
# SECTION 2 — AI ENGINE  (Ollama + Claude API)
# ═══════════════════════════════════════════════════════════

def _engine() -> str:
    return _ENGINE
def _use_claude() -> bool:  # kept for compatibility
    return _ENGINE == "claude"

# ── Ollama helpers
def _ollama_ok() -> bool:
    try: return rq.get(f"{OLLAMA_URL}/api/tags", timeout=3).status_code == 200
    except: return False

def _ollama_models() -> list:
    try: return [m["name"] for m in rq.get(f"{OLLAMA_URL}/api/tags",timeout=5).json().get("models",[])]
    except: return []

VISION_PREF = ["llava:34b","llama3.2-vision:11b","llama3.2-vision","llava:13b","llava:7b","llava"]
TEXT_PREF   = ["llama3.3:70b","llama3.1:70b","llama3.1:8b","llama3:8b","gemma3:12b","mistral","llava"]

def _pick_vision():
    ms = _ollama_models()
    for p in VISION_PREF:
        m = next((x for x in ms if p.lower() in x.lower()), None)
        if m: return m, True
    for p in TEXT_PREF:
        m = next((x for x in ms if p.lower() in x.lower()), None)
        if m: return m, False
    return (ms[0] if ms else "llava"), False

def _pick_text():
    ms = _ollama_models()
    for p in TEXT_PREF:
        m = next((x for x in ms if p.lower() in x.lower()), None)
        if m: return m
    return ms[0] if ms else "llama3"

def _ollama_vision(b64: str, prompt: str, model: str,
                   is_vis=True, timeout=900, num_predict=6000) -> str:
    TOUT = (rq.exceptions.ReadTimeout, rq.exceptions.ConnectTimeout, rq.exceptions.Timeout)
    def _call(img, include, t):
        payload = {"model":model,"prompt":prompt,"stream":False,
                   "images":[img] if (img and is_vis and include) else [],
                   "options":{"temperature":0.05,"num_predict":num_predict}}
        r = rq.post(f"{OLLAMA_URL}/api/generate", json=payload, timeout=t)
        r.raise_for_status()
        return r.json().get("response","")
    # 3-attempt progressive retry
    for attempt, (img, inc, t) in enumerate([
        (b64,   True,  timeout),
        (_compress(Image.open(io.BytesIO(base64.b64decode(b64))), 640) if b64 else b64,
         True,  int(timeout*1.5)),
        ("",    False, int(timeout*1.5)),
    ]):
        try:
            return _call(img, inc, t)
        except TOUT:
            if attempt == 2: raise RuntimeError("All vision retry attempts timed out")
        except rq.exceptions.HTTPError as e:
            if not (e.response and e.response.status_code == 500): raise
    return ""

def _ollama_text(prompt: str, model: str, timeout=1800, num_predict=8000) -> str:
    TOUT = (rq.exceptions.ReadTimeout, rq.exceptions.ConnectTimeout, rq.exceptions.Timeout)
    def _call(p, t):
        payload = {"model":model,"prompt":p,"stream":False,
                   "options":{"temperature":0.05,"num_predict":num_predict}}
        r = rq.post(f"{OLLAMA_URL}/api/generate", json=payload, timeout=t)
        r.raise_for_status()
        return r.json().get("response","")
    for p_slice, t_mult in [(prompt, 1), (prompt[:len(prompt)//2]+"\n[truncated]\n"+prompt[-6000:], 1.5)]:
        try: return _call(p_slice, int(1800*t_mult))
        except TOUT: continue
    raise RuntimeError("Text query timed out after all retries")

# ── Claude API helpers
def _claude_vision(b64: str, prompt: str) -> str:
    import anthropic
    client = anthropic.Anthropic(api_key=_CLAUDE_KEY)
    msg = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=8000,
        messages=[{"role":"user","content":[
            {"type":"image","source":{"type":"base64","media_type":"image/jpeg","data":b64}},
            {"type":"text","text":prompt}
        ]}]
    )
    return msg.content[0].text

def _claude_text(prompt: str) -> str:
    import anthropic
    client = anthropic.Anthropic(api_key=_CLAUDE_KEY)
    msg = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=8000,
        messages=[{"role":"user","content":prompt}]
    )
    return msg.content[0].text

# ── OpenAI helpers
def _openai_vision(b64: str, prompt: str) -> str:
    import openai
    client = openai.OpenAI(api_key=_OPENAI_KEY)
    resp = client.chat.completions.create(
        model="gpt-4o",
        max_tokens=8000,
        messages=[{"role":"user","content":[
            {"type":"image_url","image_url":{"url":f"data:image/jpeg;base64,{b64}","detail":"high"}},
            {"type":"text","text":prompt}
        ]}]
    )
    return resp.choices[0].message.content

def _openai_text(prompt: str) -> str:
    import openai
    client = openai.OpenAI(api_key=_OPENAI_KEY)
    resp = client.chat.completions.create(
        model="gpt-4o",
        max_tokens=8000,
        messages=[{"role":"user","content":prompt}]
    )
    return resp.choices[0].message.content

# ── Gemini helpers
GEMINI_MODEL = "gemini-1.5-flash"   # stable free-tier model

def _gemini_vision(b64: str, prompt: str) -> str:
    try:
        # Try new google-genai SDK first
        import google.genai as genai_new
        client = genai_new.Client(api_key=_GEMINI_KEY)
        img_bytes = base64.b64decode(b64)
        import google.genai.types as gtypes
        resp = client.models.generate_content(
            model=GEMINI_MODEL,
            contents=[prompt, gtypes.Part.from_bytes(data=img_bytes, mime_type="image/jpeg")])
        return resp.text
    except Exception:
        pass
    # Fallback: legacy google-generativeai SDK
    import google.generativeai as genai
    genai.configure(api_key=_GEMINI_KEY)
    model = genai.GenerativeModel(GEMINI_MODEL)
    img = Image.open(io.BytesIO(base64.b64decode(b64)))
    resp = model.generate_content([prompt, img])
    return resp.text

def _gemini_text(prompt: str) -> str:
    try:
        import google.genai as genai_new
        client = genai_new.Client(api_key=_GEMINI_KEY)
        resp = client.models.generate_content(model=GEMINI_MODEL, contents=prompt)
        return resp.text
    except Exception:
        pass
    import google.generativeai as genai
    genai.configure(api_key=_GEMINI_KEY)
    model = genai.GenerativeModel(GEMINI_MODEL)
    resp = model.generate_content(prompt)
    return resp.text

# ── Unified interface
def ai_vision(b64:str, prompt:str, vm="", is_vis=True, timeout=900, num_predict=6000) -> str:
    e = _engine()
    if e == "claude":  return _claude_vision(b64, prompt)
    if e == "openai":  return _openai_vision(b64, prompt)
    if e == "gemini":  return _gemini_vision(b64, prompt)
    return _ollama_vision(b64, prompt, vm, is_vis, timeout, num_predict)

def ai_text(prompt:str, tm="", timeout=1800, num_predict=8000) -> str:
    e = _engine()
    if e == "claude":  return _claude_text(prompt)
    if e == "openai":  return _openai_text(prompt)
    if e == "gemini":  return _gemini_text(prompt)
    return _ollama_text(prompt, tm, timeout, num_predict)

# ═══════════════════════════════════════════════════════════
# SECTION 3 — IMAGE HELPERS
# ═══════════════════════════════════════════════════════════

def _compress(img: Image.Image, max_dim=768) -> str:
    if img.mode not in ("RGB","L"): img = img.convert("RGB")
    w, h = img.size
    if max(w,h) > max_dim:
        s = max_dim / max(w,h)
        img = img.resize((int(w*s), int(h*s)), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, "JPEG", quality=82)
    return base64.b64encode(buf.getvalue()).decode()

def pdf_to_images(path: Path, max_pages=5, max_dim=768) -> list[dict]:
    """Return list of {b64, text, page}"""
    doc = fitz.open(str(path)); out = []
    for i in range(min(len(doc), max_pages)):
        page = doc[i]
        text = page.get_text("text")
        pix  = page.get_pixmap(matrix=fitz.Matrix(1.5, 1.5))
        img  = Image.frombytes("RGB",[pix.width,pix.height],pix.samples)
        out.append({"b64":_compress(img,max_dim),"text":text[:5000],"page":i+1})
    doc.close(); return out

def pdf_extract_text(path: Path, max_pages=6) -> str:
    try:
        doc = fitz.open(str(path))
        parts = [f"[Page {i+1}]\n{doc[i].get_text('text')}"
                 for i in range(min(len(doc),max_pages))
                 if doc[i].get_text("text").strip()]
        doc.close()
        return "\n\n".join(parts)[:12000]
    except: return ""

# ═══════════════════════════════════════════════════════════
# SECTION 4 — BBS PARSER
# ═══════════════════════════════════════════════════════════

def _normalise_header(h: str) -> str:
    return str(h).upper().replace(" ","").replace("_","").replace(".","").strip()

HEADER_MAP = {
    "BARMARK":"bar_mark","MARK":"bar_mark","BARNO":"bar_mark","NO":"bar_mark",
    "MEMBER":"member","DESCRIPTION":"member","ELEMENT":"member",
    "DIA":"dia","DIAMETER":"dia","SIZE":"dia","BARDIA":"dia",
    "SHAPECODE":"shape_code","SHAPE":"shape_code","CODE":"shape_code",
    "TOTALLENGTH":"total_len","TOTALLENGTHCUTTING":"total_len",
    "TOTAL":"total_len","LENGTH":"total_len","CUTTINGLENGTH":"total_len",
    "QTY":"qty","QUANTITY":"qty","NUMBER":"qty","NO":"qty","NOOF":"qty",
    "A":"leg_A","B":"leg_B","C":"leg_C","D":"leg_D","E":"leg_E","F":"leg_F",
    "LEGA":"leg_A","LEGB":"leg_B","LEGC":"leg_C","LEGD":"leg_D",
}

def parse_bbs_excel(path: Path) -> list[dict]:
    """Parse BBS directly from Excel — deterministic, no AI, handles multi-sheet."""
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    all_rows = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        col_map: dict[str,int] = {}
        data_start = 0
        # Find header row (scan first 30 rows)
        for ri, row in enumerate(ws.iter_rows(max_row=30, values_only=True)):
            hits = sum(1 for c in row if c and _normalise_header(str(c)) in HEADER_MAP)
            if hits >= 2:
                for ci, cell in enumerate(row):
                    if cell:
                        key = HEADER_MAP.get(_normalise_header(str(cell)))
                        if key and key not in col_map:
                            col_map[key] = ci
                data_start = ri + 1
                break
        if not col_map: continue

        def _get(row_vals, key):
            ci = col_map.get(key)
            return row_vals[ci] if ci is not None and ci < len(row_vals) else None

        for row in ws.iter_rows(min_row=data_start+1, values_only=True):
            if not any(row): continue
            mark = _get(row, "bar_mark")
            if not mark or str(mark).strip() in ("","None","-","N/A"): continue
            def _mm(v):
                try: return round(float(str(v).replace(",","")))
                except: return None
            record = {
                "bar_mark": str(mark).strip(),
                "member":   str(_get(row,"member") or "").strip(),
                "dia":      str(_get(row,"dia") or "").strip(),
                "shape_code": str(_get(row,"shape_code") or "").strip(),
                "total_len":  _mm(_get(row,"total_len")),
                "qty":        _mm(_get(row,"qty")),
                "legs": {L: _mm(_get(row,f"leg_{L}")) for L in "ABCDEF"},
            }
            all_rows.append(record)

    # Dedup exact rows
    seen = set(); deduped = []
    for r in all_rows:
        k = f"{r['bar_mark']}|{r['member']}|{r['qty']}"
        if k not in seen: seen.add(k); deduped.append(r)
    return deduped

def parse_bbs_ai(path: Path, vm: str, tm: str, is_vis: bool, push_fn) -> list[dict]:
    """Fallback BBS parser using AI (for scanned PDFs / images)."""
    PROMPT = """Extract every bar mark from this Bar Bending Schedule. Return ONLY valid JSON:
{"rows":[{"bar_mark":"","member":"","dia":"","shape_code":"","total_len":0,"qty":0,
"legs":{"A":null,"B":null,"C":null,"D":null,"E":null,"F":null}}]}
List EVERY row. Do NOT merge rows with same mark."""
    rows = []
    ext = path.suffix.lower()
    if ext == ".pdf":
        text = pdf_extract_text(path, max_pages=20)
        if len(text.strip()) > 300:
            # Text-layer PDF
            BATCH = 5000
            batches = [text[i:i+BATCH] for i in range(0, len(text), BATCH)]
            for bi, batch in enumerate(batches):
                push_fn(f"  BBS text batch {bi+1}/{len(batches)}...")
                try:
                    raw = ai_text(PROMPT + f"\n\nBBS DATA:\n{batch}", tm, timeout=1800, num_predict=6000)
                    obj,_ = _parse_json(raw)
                    if obj: rows.extend(obj.get("rows",[]))
                except Exception as e:
                    push_fn(f"  ⚠ Batch {bi+1}: {e}")
        else:
            # Scanned PDF
            imgs = pdf_to_images(path, max_pages=10, max_dim=640)
            for i, img in enumerate(imgs):
                push_fn(f"  BBS vision page {i+1}/{len(imgs)}...")
                try:
                    raw = ai_vision(img["b64"], PROMPT, vm, is_vis, timeout=1200, num_predict=4000)
                    obj,_ = _parse_json(raw)
                    if obj: rows.extend(obj.get("rows",[]))
                except Exception as e:
                    push_fn(f"  ⚠ Page {i+1}: {e}")
    elif ext in (".png",".jpg",".jpeg",".tiff",".bmp"):
        img = Image.open(path)
        try:
            raw = ai_vision(_compress(img,640), PROMPT, vm, is_vis, timeout=1200, num_predict=4000)
            obj,_ = _parse_json(raw)
            if obj: rows.extend(obj.get("rows",[]))
        except Exception as e:
            push_fn(f"  ⚠ Vision: {e}")
    return rows

# ═══════════════════════════════════════════════════════════
# SECTION 5 — DRAWING PARSER
# ═══════════════════════════════════════════════════════════

DRAW_PROMPT = """You are a senior rebar detailer reading a rebar shop drawing for QA/QC.
Extract precisely:

1. EVERY bar label with FULL text (e.g. "4T16-01-150", "2x3T20-05-200 EF")
   For each: bar_mark, dia, spacing_mm, qty, multipliers, face_notation(EF/ES/EE/T&B etc.)

2. EVERY section cut (A-A, B-B etc.) showing:
   - Cover callout present? (e.g. "35 CLR") — exact value
   - Concrete dimension shown? (e.g. "300x600") — exact value
   - Concrete dim MISSING? State explicitly.

4. DIMENSIONS: spans, depths, widths with exact values and units

5. GENERAL NOTES: concrete grade, cover specification, standard

6. TITLE BLOCK: drawing number, revision, project name

7. CALCULATIONS shown on drawing (e.g. "12d=192mm", "1.3L=4160mm")

8. ELEMENT MARKS visible (B1, C1, S1 etc.) with floor/level if shown

State MISSING explicitly. Write exact numbers only."""

REF_PROMPT = """You are reading a STRUCTURAL REINFORCEMENT PLAN (reference/IFC drawing).
Extract ALL rebar information — this drawing IS the reference standard.

For EVERY bar visible, extract:
- bar_mark, dia (e.g. T16), spacing_mm (e.g. 150), count (if shown as qty not spacing)
- location: top/bottom/additional_top/additional_bottom/EF/ES
- direction: longitudinal/transverse/diagonal
- range: extent shown (e.g. "full span", "support zone 1200mm from col face")
- full callout text exactly as shown

Also extract:
- All dimensions visible (spans, depths)
- Concrete grade, cover spec
- Element marks (S1, B1 etc.)
- Grid lines visible

State MISSING if something expected is absent."""

def describe_drawing(path: Path, audit_mode: str, vm: str, is_vis: bool,
                     prompt: str, push_fn) -> str:
    """Extract text + optional vision description from a drawing PDF/image."""
    ext = path.suffix.lower()
    parts = []
    text = ""

    # Always extract PDF text first
    if ext == ".pdf":
        text = pdf_extract_text(path, max_pages=6)
        if text.strip():
            parts.append(f"[PDF TEXT EXTRACTION]\n{text}")
            push_fn(f"  PDF text: {len(text)} chars extracted")

    # Vision based on mode
    vis_tiles = {"fast":0, "smart":1, "deep":4}.get(audit_mode, 1)
    if len(text.strip()) < 400:
        vis_tiles = max(vis_tiles, 1)  # force vision for scanned pages

    if is_vis and vis_tiles > 0:
        if ext == ".pdf":
            imgs = pdf_to_images(path, max_pages=min(vis_tiles,4), max_dim=768)
        else:
            img = Image.open(path)
            if img.mode not in ("RGB","L"): img = img.convert("RGB")
            imgs = [{"b64":_compress(img,768),"text":"","page":1}]

        for i, img_d in enumerate(imgs[:vis_tiles]):
            push_fn(f"  Vision tile {i+1}/{min(vis_tiles,len(imgs))}...")
            try:
                desc = ai_vision(img_d["b64"], prompt, vm, is_vis, timeout=1800, num_predict=6000)
                parts.append(f"[VISION TILE {i+1}]\n{desc}")
                push_fn(f"  ✓ Tile {i+1}: {len(desc)} chars")
            except Exception as e:
                push_fn(f"  ⚠ Tile {i+1}: {e}")
                if img_d.get("text"):
                    parts.append(f"[TILE {i+1} TEXT FALLBACK]\n{img_d['text']}")

    return "\n\n".join(parts) if parts else "No content extracted."

# ═══════════════════════════════════════════════════════════
# SECTION 6 — JSON HELPERS
# ═══════════════════════════════════════════════════════════

def _parse_json(raw: str) -> tuple:
    try:
        s = raw.replace("```json","").replace("```","").strip()
        si = min((s.find("{") if "{" in s else len(s)),
                 (s.find("[") if "[" in s else len(s)))
        if si == len(s): return None, "No JSON"
        closer = "}" if s[si]=="{" else "]"
        ei = s.rfind(closer)+1
        return json.loads(s[si:ei]), None
    except json.JSONDecodeError as ex:
        return None, str(ex)

# ═══════════════════════════════════════════════════════════
# SECTION 7 — AUDIT CHECK PROMPTS  (4 focused prompts)
# ═══════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════
# SECTION 7 — MODULAR PROMPT BUILDERS
# ═══════════════════════════════════════════════════════════

SCHEMA_BASE = ('{"checks":[{"check":"","bar_mark":"","element":"","section_ref":"",'
               '"severity":"error","issue":"","fix":"","confidence":"High",'
               '"element_type":"General","element_id":"",'
               '"bbox":{"x":0.1,"y":0.1,"w":0.15,"h":0.08}}]}')

BBOX_GUIDE = """
BBOX COORDINATES: estimate normalized position (0.0–1.0) of the issue on the drawing.
  Sections/elevation → right half (x≈0.55–0.90). Plan view → left/centre (x≈0.02–0.50).
  Title block → bottom-right (x≈0.70, y≈0.85). General notes → bottom-left (x≈0.02, y≈0.80).
  Never use x=0,y=0 unless the error is literally at the top-left corner."""

def _build_drawing_prompt(shop_desc:str, std:str, rules:dict, enabled:set) -> str:
    cb = rules.get("r_cover_beam",""); cs = rules.get("r_cover_slab","")
    cc = rules.get("r_cover_col",""); bs = rules.get("r_bar_sizes","")
    cover_str = f"Required covers — Beam:{cb}mm Slab:{cs}mm Column:{cc}mm." if any([cb,cs,cc]) else ""
    sizes_str = f"Permitted bar sizes: {bs}." if bs else ""

    sections = []
    if "general_arrangements" in enabled:
        sections.append(f"""CHECK: general_arrangements
  A) cover_missing — Every section cut MUST show cover callout (e.g. "35 CLR", "40mm cover").
     Missing or wrong cover = error. {cover_str}
  B) concrete_dim_missing — Every section MUST show concrete member size:
     Beams: width×depth (300×600). Slabs: thickness (200THK). Columns: width×depth. Walls: thickness.
     Section shows rebar but no concrete size = critical error.
  C) missing_callout — Warning if missing: drawing number, revision, project name,
     concrete grade, cover note, steel grade, scale/NTS, at least one section per element type.
  D) spell_error — Flag any obvious spelling mistake in labels, notes, title block.
  E) calc_error — Verify every formula: nd→n×dia_mm, Factor×L→factor×L_value, 0.8x→0.8×x,
     arithmetic sums. Tolerance ±1mm. Wrong = critical error.""")

    if "label_multiplier" in enabled:
        sections.append(f"""CHECK: label_multiplier
  Bar label format: [qty_prefix][dia][mark][spacing] [face_notations_at_end]
  Count "x" or "×" characters BEFORE the bar dia letter (T/R/Y/H/B):
    0 multipliers → "4T16-01-150"         → 0 notations required  ✓
    1 multiplier  → "2×4T16-01-150"        → EXACTLY 1 notation   e.g. "EF" ✓
    2 multipliers → "2×2×4T16-01-150"      → EXACTLY 2 notations  e.g. "EF ES" ✓
  Valid notations: EF ES EE T&B T1&B1 T2&B2 NF FF B1 T1
  Wrong count = error. State exact label text and what is missing/extra.
  {f"Only these bar sizes permitted: {bs}" if bs else ""}""")

    if "lap_qty" in enabled:
        sections.append("""CHECK: lap_qty
  Lapping bars MUST have the same quantity.
  If bar mark A (qty N) is shown lapping with bar mark B, bar B must also have qty N.
  Method: look for bars drawn overlapping or notes like "lap with bar 05".
  Same member, adjacent levels in column = lapping relationship.
  Qty mismatch between lapping bars = error. Exception: opening reduces range (note it).""")

    if "missing_bar" in enabled:
        sections.append("""CHECK: missing_bar
  A) missing_bar — Every concrete section shown MUST have rebar within it.
     A concrete section/element with NO bars drawn inside = error.
  B) extra_bar — Bars must not be shown at non-structural locations (voids, openings, expansion joints).
     Bar shown at an opening or non-concrete area = error.""")

    if not sections:
        return '{"checks":[]}'

    return f"""Senior rebar QA/QC audit. Standard: {std}. {cover_str} {sizes_str}
Return ONLY valid JSON. One entry per issue. Return {{"checks":[]}} if no issues.
{BBOX_GUIDE}

SHOP DRAWING:
{shop_desc[:5000]}

{"=" * 50}
{"".join(f"{chr(10)}{s}{chr(10)}" for s in sections)}
{"=" * 50}
{SCHEMA_BASE}"""


def _build_bbs_prompt(bbs_rows:list, shop_desc:str, std:str, enabled:set) -> str:
    bbs_json = __import__('json').dumps(bbs_rows, indent=None)[:5000]
    sections = []
    if "bbs_check" in enabled:
        sections.append("""CHECK: bbs_check
  A) bbs_qty — Sum drawing label quantities per bar mark (only labels WITH a number prefix,
     e.g. "4T16-01"=qty 4; ignore section references like "-01-").
     Compare to BBS total qty (sum all BBS rows for same mark). Mismatch = error.
  B) bbs_shape — Verify shape code matches required legs A–F:
     00=A, 11=A+B, 13=A+B+C, 21=A+B, 25=A+B+C, 26=A, 41=A+B. Missing leg = error.
  C) bbs_transport — Any bar total_len > 12000mm = critical error.
  D) bbs_rounding — Any BBS leg dimension NOT ending in 00/25/50/75 = warning.
  E) member_mark — Same bar mark with DIFFERENT dia or shape_code across BBS rows = error.""")
    if "bar_length" in enabled:
        sections.append("""CHECK: bar_length
  For each bar mark where a concrete dimension AND cover is visible in the drawing:
  Expected bar length = concrete_dim - 2 × cover (tolerance ±10mm).
  Compare to BBS total_len. Mismatch = error.
  State: concrete_dim used, cover used, expected length, BBS length.""")
    if not sections: return '{"checks":[]}'
    return f"""Senior rebar QA/QC. Standard: {std}. Return ONLY valid JSON.
{BBOX_GUIDE}

BBS DATA:
{bbs_json}

SHOP DRAWING:
{shop_desc[:3000]}

{"=" * 50}
{"".join(f"{chr(10)}{s}{chr(10)}" for s in sections)}
{"=" * 50}
{SCHEMA_BASE}"""


def _build_ref_prompt(ref_desc:str, shop_desc:str, std:str, rules:dict, enabled:set) -> str:
    cb = rules.get("r_cover_beam",""); bs = rules.get("r_bar_sizes","")
    sections = []
    if "bar_dia" in enabled:
        sections.append(f"""CHECK: bar_dia
  For EVERY bar in the reference plan, find the same bar in the shop drawing.
  Compare bar DIAMETER (T10/T12/T16/T20/T25/T32 etc.).
  Mismatch = CRITICAL error. {f"Permitted sizes: {bs}." if bs else ""}
  Bar in reference NOT found in shop = CRITICAL error.
  Extra bar in shop NOT in reference = CRITICAL error (unauthorised).""")
    if "spacing" in enabled:
        sections.append("""CHECK: spacing
  For EVERY bar in the reference plan, compare SPACING in shop drawing.
  Reference spacing (e.g. @150mm) must match shop drawing exactly.
  Mismatch = CRITICAL error.
  Also check: bar LOCATION (top/bottom/additional) must match reference.""")
    if not sections: return '{"checks":[]}'
    return f"""Senior rebar QA/QC. Standard: {std}. Return ONLY valid JSON.
{BBOX_GUIDE}

REFERENCE PLAN (gold standard):
{ref_desc[:6000]}

SHOP DRAWING:
{shop_desc[:4000]}

{"=" * 50}
{"".join(f"{chr(10)}{s}{chr(10)}" for s in sections)}
{"=" * 50}
{SCHEMA_BASE}"""


def _drawing_quality_prompt(shop_desc: str, std: str, rules: dict) -> str:
    cover_beam = rules.get("r_cover_beam","")
    cover_slab = rules.get("r_cover_slab","")
    cover_col  = rules.get("r_cover_col","")
    bar_sizes  = rules.get("r_bar_sizes","")
    max_link   = rules.get("r_max_link_spacing","")

    cover_text = ""
    if cover_beam: cover_text += f"Beam cover: {cover_beam}mm. "
    if cover_slab: cover_text += f"Slab cover: {cover_slab}mm. "
    if cover_col:  cover_text += f"Column cover: {cover_col}mm. "

    rules_text = ""
    if bar_sizes:  rules_text += f"Permitted bar sizes: {bar_sizes}. "
    if max_link:   rules_text += f"Max link spacing: {max_link}mm. "

    schema = ('{"checks":['
        '{"check":"","bar_mark":"","element":"","section_ref":"",'
        '"severity":"error","issue":"","fix":"","confidence":"High",'
        '"element_type":"General","element_id":"",'
        '"bbox":{"x":0.0,"y":0.0,"w":0.15,"h":0.08}}]}')
    return f"""You are a senior rebar detailer doing QA/QC. Standard: {std}.
{cover_text}{rules_text}

SHOP DRAWING CONTENT:
{shop_desc[:5000]}

Read carefully. Return one JSON entry per issue. Return {{"checks":[]}} if no issues found.

BBOX RULE: Estimate the normalized position (0.0–1.0) of each issue on the drawing:
  - Top-left area = low x, low y. Bottom-right = high x, high y.
  - Title block is usually bottom-right: x≈0.7, y≈0.85
  - Sections are usually right half: x≈0.5–0.9
  - Plan view is usually left/center: x≈0.0–0.5
  - Never use x=0,y=0 unless the error is literally at the top-left corner.

CHECK A — COVER CALLOUTS (check="cover_missing"):
  Every section (A-A, B-B etc.) MUST show cover/clear callout e.g. "35 CLR", "40mm CLR".
  Missing = error. Cover value < required = error.
  {f"Required: Beams {cover_beam}mm, Slabs {cover_slab}mm, Columns {cover_col}mm." if cover_beam else "Check against code minimum for the standard."}

CHECK B — CONCRETE MEMBER DIMENSIONS (check="concrete_dim_missing"):
  Every section MUST show the concrete size explicitly:
  Beams → width×depth e.g. 300×600. Slabs → thickness e.g. 200THK.
  Columns → width×depth e.g. 400×400. Walls → thickness e.g. 200mm.
  Section shows rebar but NO concrete size → severity=error.

CHECK C — LABEL MULTIPLIER NOTATION (check="label_multiplier"):
  Bar label format: [qty_prefix][dia][mark][spacing] [face_notations]
  The qty_prefix uses "x" or "×" as multiplier separator:
    No multiplier  → "4T16-01-150"            → 0 face notations required  ✓
    One multiplier → "2×4T16-01-150"           → EXACTLY 1 notation at end  e.g. "EF"
    Two multipliers→ "2×2×4T16-01-150"         → EXACTLY 2 notations at end e.g. "EF ES"
  Valid notations: EF, ES, EE, T&B, T1&B1, T2&B2, NF, FF, B1, T1
  Count "x" or "×" symbols BEFORE the first letter T/R/Y/H/B to get multiplier count.
  Wrong notation count → severity=error. Note exact label found and what is wrong.
  {f"Only these bar sizes are permitted: {bar_sizes}" if bar_sizes else ""}

CHECK D — DRAWING COMPLETENESS (check="missing_callout"):
  Warning if missing: drawing number, revision, project name, concrete grade, cover note,
  steel grade/standard, scale or NTS, at least one section per element type.

CHECK E — CALCULATION ERRORS (check="calc_error"):
  Find EVERY formula on drawing e.g. "12d=192", "1.3L=4160(L=3200)", "0.8×3000=2400".
  Compute correctly: nd → n×dia_mm. Factor×L → factor×L_value. 0.8x → 0.8×x.
  Tolerance ±1mm. Wrong answer → severity=error.

CHECK F — BAR DIAMETER vs RULES (check="dia_vs_rules"):
  {f"Permitted bar sizes are: {bar_sizes}. Flag any bar mark using a non-permitted size." if bar_sizes else "Skip this check — no bar size rules set."}
  {f"Flag any link/stirrup spacing > {max_link}mm." if max_link else ""}

{schema}"""

def _bbs_check_prompt(bbs_rows: list, shop_desc: str, std: str) -> str:
    bbs_json = json.dumps(bbs_rows, indent=None)[:5000]
    schema = ('{"checks":['
        '{"check":"bbs_qty","bar_mark":"","bbs_qty":0,"drawing_qty":0,"match":true,'
        '"severity":"error","issue":"","fix":"","element_type":"General","element_id":"",'
        '"confidence":"High","bbox":{"x":0,"y":0,"w":0.1,"h":0.05}}]}')
    return f"""Senior structural engineer QA/QC audit. Standard: {std}.
Return ONLY valid JSON.

BBS DATA:
{bbs_json}

SHOP DRAWING DESCRIPTION:
{shop_desc[:4000]}

RUN THESE CHECKS — return one entry per issue found:

A) bbs_qty: For each bar mark, sum all drawing label quantities (only labels WITH a number prefix,
   e.g. "4T16-01" = qty 4; ignore references like "-01-" at sections).
   Compare sum to BBS qty (sum all BBS rows for same mark). Mismatch = error.

B) bbs_length: For bars where a concrete dimension and cover is visible:
   check concrete_dim - 2*cover ≈ BBS total_len (±10mm). Mismatch = error.

C) bbs_shape: For each bar mark, verify shape code matches required legs (A-F).
   Shape 00=A only, 11=A,B, 13=A,B,C, 21=A,B, 25=A,B,C, 41=A,B. Missing leg = error.

D) bbs_transport: Flag any bar where total_len > 12000mm. Critical error.

E) bbs_rounding: Flag any BBS leg dimension NOT ending in 00, 25, 50, or 75. Warning.

F) member_mark: If same bar mark appears in multiple BBS rows with DIFFERENT dia or shape_code = error.

{schema}"""

def _cover_dim_prompt(shop_desc: str, std: str, rules: dict) -> str:
    cover_rule = rules.get("cover","")
    schema = ('{"checks":['
        '{"check":"cover_missing","element":"","section_ref":"","cover_shown":"",'
        '"concrete_dim_shown":"","severity":"error","issue":"","fix":"",'
        '"element_type":"General","element_id":"","confidence":"High",'
        '"bbox":{"x":0,"y":0,"w":0.1,"h":0.05}}]}')
    return f"""Senior structural engineer QA/QC. Standard: {std}.
{f"Required cover: {cover_rule}" if cover_rule else ""}
Return ONLY valid JSON.

SHOP DRAWING DESCRIPTION:
{shop_desc[:10000]}

CHECK EVERY section cut, plan, and elevation for:

A) cover_missing: Is a cover/clear callout shown at this section?
   (e.g. "35 CLR", "40mm cover"). MISSING = error.

B) concrete_dim_missing: Is the concrete member size shown?
   Beams: width x depth (e.g. 300x600). Slabs: thickness (e.g. 200 THK).
   Columns: width x depth. MISSING = critical error.

C) label_multiplier: Bar labels with ONE x multiplier (e.g. 2x4T16-01-150)
   MUST have exactly ONE face notation at end: EF, ES, EE, T&B, T1&B1, or T2&B2.
   Labels with TWO x multipliers MUST have exactly TWO notations.
   Wrong count = error.

D) rebar_boundary: Any bar body crossing OUTSIDE the concrete outline = error.

Return one entry per issue found. Return {{"checks":[]}} if no issues.
{schema}"""

def _plan_vs_shop_prompt(ref_desc: str, shop_desc: str, std: str, rules: dict = None) -> str:
    rules = rules or {}
    schema = ('{"checks":['
        '{"check":"plan_vs_shop","bar_mark":"","dia_ref":"","dia_shop":"",'
        '"spacing_ref":"","spacing_shop":"","location_ref":"","severity":"error",'
        '"issue":"","fix":"","element_type":"General","element_id":"","confidence":"High",'
        '"bbox":{"x":0,"y":0,"w":0.1,"h":0.05}}]}')
    rules_txt = ""
    cb = rules.get("r_cover_beam",""); cs = rules.get("r_cover_slab",""); bs = rules.get("r_bar_sizes","")
    if cb: rules_txt += f"Required beam cover: {cb}mm. "
    if cs: rules_txt += f"Required slab cover: {cs}mm. "
    if bs: rules_txt += f"Permitted bar sizes: {bs}. "
    return f"""Senior structural engineer QA/QC. Standard: {std}. Return ONLY valid JSON.
{rules_txt}

REFERENCE PLAN (standard — must be matched exactly):
{ref_desc[:8000]}

SHOP DRAWING (being checked):
{shop_desc[:8000]}

CHECK: For every bar in the reference plan, verify the shop drawing shows:
  A) Same bar DIAMETER (T16, T20 etc.) — mismatch = CRITICAL error
  B) Same SPACING (e.g. @150mm) — mismatch = CRITICAL error
  C) Same LOCATION (top/bottom/additional top etc.) — mismatch = error
  D) Bar EXISTS in shop drawing — missing = CRITICAL error
  E) Any EXTRA bar in shop drawing NOT in reference plan = CRITICAL error (unauthorised)
  F) Bar MARK difference only (same dia+spacing) = WARNING only

Return one entry per mismatch. Return {{"checks":[]}} if reference not provided or all match.
{schema}"""

def _calc_check_prompt(shop_desc: str, std: str) -> str:
    schema = ('{"checks":['
        '{"check":"calc_error","formula":"","inputs":{},"stated_result":0,"computed_result":0,'
        '"match":true,"severity":"error","issue":"","fix":"","element_type":"General",'
        '"element_id":"","confidence":"High","bbox":{"x":0,"y":0,"w":0.1,"h":0.05}}]}')
    return f"""Senior structural engineer QA/QC. Standard: {std}. Return ONLY valid JSON.

SHOP DRAWING DESCRIPTION:
{shop_desc[:10000]}

VERIFY every formula/calculation shown on the drawing:

A) nd or nphi: bar dia multiples (e.g. "12d=192" with d=16 → 12x16=192 ✓)
   If bar dia is T16 then d=16. If T20 then d=20.

B) Factor x L: span fractions (e.g. "1.3L=4160" with L=3200 → 1.3x3200=4160 ✓)
   Find L from nearby dimension callouts.

C) 0.8x: 80% of value (e.g. "0.8x3000=2400" → 0.8x3000=2400 ✓)

D) Arithmetic: sums of dims (e.g. "300+150+40d=1090" with d=16 → 300+150+640=1090 ✓)

For each formula found:
  - Extract the stated result from the drawing
  - Compute the correct result using stated inputs
  - If stated_result ≠ computed_result (>1mm tolerance) = CRITICAL error

Return {{"checks":[]}} if no formulas found or all correct.
{schema}"""

# ═══════════════════════════════════════════════════════════
# SECTION 8 — DETERMINISTIC CHECKS (no AI)
# ═══════════════════════════════════════════════════════════

def deterministic_bbs_checks(bbs_rows: list) -> list:
    """Fast, reliable checks that don't need AI."""
    issues = []
    from collections import defaultdict

    # Transport length
    for r in bbs_rows:
        tl = r.get("total_len") or 0
        if tl and tl > MAX_BAR_LEN_MM:
            issues.append({"check":"bbs_transport","bar_mark":r["bar_mark"],
                "member":r.get("member",""), "total_len":tl,
                "excess_mm":tl-MAX_BAR_LEN_MM,
                "severity":"error","confidence":"High",
                "issue":f"Bar {r['bar_mark']} total length {tl}mm exceeds {MAX_BAR_LEN_MM}mm",
                "fix":f"Split bar with coupler or lap splice to ≤{MAX_BAR_LEN_MM}mm",
                "element_type":"General","element_id":r["bar_mark"],
                "bbox":{"x":0,"y":0,"w":0.08,"h":0.04}})

    # Rounding (25mm multiples)
    for r in bbs_rows:
        for leg, val in (r.get("legs") or {}).items():
            if val is None: continue
            try:
                v = float(val)
                if v > 0 and round(v % 100, 1) not in (0.0, 25.0, 50.0, 75.0):
                    nearest = round(round(v/25)*25)
                    issues.append({"check":"bbs_rounding","bar_mark":r["bar_mark"],
                        "member":r.get("member",""),"leg":leg,"value_mm":v,
                        "nearest_valid_mm":nearest,
                        "severity":"warning","confidence":"High",
                        "issue":f"Bar {r['bar_mark']} Leg {leg}={v}mm — not a 25mm multiple",
                        "fix":f"Round to {nearest}mm",
                        "element_type":"General","element_id":r["bar_mark"],
                        "bbox":{"x":0,"y":0,"w":0.08,"h":0.04}})
            except: pass

    # Member mark consistency
    by_mark = defaultdict(list)
    for r in bbs_rows:
        if r.get("bar_mark"): by_mark[r["bar_mark"]].append(r)
    for mark, rows in by_mark.items():
        if len(rows) < 2: continue
        dias   = {str(r.get("dia","")).strip() for r in rows} - {"","None"}
        shapes = {str(r.get("shape_code","")).strip() for r in rows} - {"","None"}
        if len(dias) > 1 or len(shapes) > 1:
            members = [r.get("member","?") for r in rows]
            bad = []
            if len(dias)>1:   bad.append(f"dia: {', '.join(sorted(dias))}")
            if len(shapes)>1: bad.append(f"shape_code: {', '.join(sorted(shapes))}")
            issues.append({"check":"member_mark","bar_mark":mark,
                "members":members,"inconsistent_fields":bad,
                "severity":"error","confidence":"High",
                "issue":f"Bar {mark} has conflicting {'; '.join(bad)} across members {members[:5]}",
                "fix":"Assign unique bar mark for each different bar specification",
                "element_type":"General","element_id":mark,
                "bbox":{"x":0,"y":0,"w":0.08,"h":0.04}})

    return issues

# ═══════════════════════════════════════════════════════════
# SECTION 9 — ANNOTATION
# ═══════════════════════════════════════════════════════════

SEV_COL = {"error":(220,38,38),"critical":(220,38,38),
           "warning":(234,115,0),"info":(202,170,0)}

def annotate(b64: str, issues: list) -> str:
    """Draw numbered markers on the drawing at each issue's bbox."""
    raw = Image.open(io.BytesIO(base64.b64decode(b64)))
    W0, H0 = raw.size
    S = 2
    img = raw.convert("RGBA").resize((W0*S, H0*S), Image.LANCZOS)
    W, H = img.size
    ov  = Image.new("RGBA",(W,H),(0,0,0,0))
    draw= ImageDraw.Draw(ov)

    # Load fonts
    bold = None
    for fp in ["C:/Windows/Fonts/arialbd.ttf","/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"]:
        try: bold = ImageFont.truetype(fp, max(14,W//55)); break
        except: pass
    if not bold: bold = ImageFont.load_default()

    placed = []
    valid  = [x for x in issues if x.get("bbox") and
              (x["bbox"].get("w",0)>0 or x["bbox"].get("h",0)>0)]

    for idx, issue in enumerate(valid):
        bb  = issue["bbox"]
        col = SEV_COL.get((issue.get("severity") or "warning").lower(), SEV_COL["warning"])
        x1  = max(0, int(bb.get("x",0)*W))
        y1  = max(0, int(bb.get("y",0)*H))
        x2  = min(W, x1+max(24, int(bb.get("w",0.08)*W)))
        y2  = min(H, y1+max(18, int(bb.get("h",0.04)*H)))

        # Highlight box
        draw.rectangle([x1,y1,x2,y2], fill=col+(35,), outline=col+(210,), width=max(2,W//300))

        # Circle placement with stagger
        cr  = max(14, W//55)
        ecx = (x1+x2)//2; ecy = (y1+y2)//2
        cx_c= ecx; cy_c = max(cr+4, y1-cr-8)
        for px,py in placed:
            if abs(cx_c-px)<2*cr+8 and abs(cy_c-py)<2*cr+8:
                cx_c += 2*cr+10
                if cx_c+cr > W-4: cx_c=max(cr+4,x1); cy_c+=2*cr+10
        placed.append((cx_c,cy_c))

        # Pointer dot + leader
        draw.ellipse([ecx-5,ecy-5,ecx+5,ecy+5], fill=col+(255,))
        draw.line([(cx_c,cy_c),(ecx,ecy)], fill=col+(190,), width=max(2,W//400))

        # Numbered circle
        draw.ellipse([cx_c-cr,cy_c-cr,cx_c+cr,cy_c+cr],
                     fill=(255,255,255,230), outline=col+(255,), width=max(2,W//350))
        draw.ellipse([cx_c-cr+2,cy_c-cr+2,cx_c+cr-2,cy_c+cr-2], fill=col+(220,))
        draw.text((cx_c,cy_c), str(idx+1), fill=(255,255,255,255), font=bold, anchor="mm")

    out = Image.alpha_composite(img, ov).convert("RGB").resize((W0,H0),Image.LANCZOS)
    buf = io.BytesIO(); out.save(buf,"PNG",optimize=True)
    return base64.b64encode(buf.getvalue()).decode()

# ═══════════════════════════════════════════════════════════
# SECTION 10 — JOB RUNNER
# ═══════════════════════════════════════════════════════════

def push(jid, etype, data):
    with LOCK:
        if jid in JOBS: JOBS[jid]["events"].append({"type":etype,"data":data})

def log(jid, msg, pct=None):
    d = {"msg":msg}
    if pct is not None: d["pct"] = pct
    push(jid,"progress",d)

def cancelled(jid) -> bool:
    with LOCK: return JOBS.get(jid,{}).get("cancelled",False)

def run_job(jid, shop_path, ref_path, bbs_path, conditions, shop_name):
    try:
        audit_mode = conditions.get("audit_mode","smart").lower()
        std        = conditions.get("standard","BS8666")
        print("\n"+"="*55+f"\n[{jid}] JOB START | mode={audit_mode} | std={std}\n"+"="*55, flush=True)
        rules      = {k:conditions.get(k,"") for k in ("r_cover_beam","r_cover_slab","r_cover_col","r_bar_sizes","r_max_link_spacing","custom_notes")}

        # Parse enabled checks (comma-separated; default = all enabled)
        ALL_CHECKS = {"label_multiplier","bar_dia","spacing","general_arrangements",
                      "lap_qty","bbs_check","bar_length","missing_bar"}
        raw_ec = conditions.get("enabled_checks","").strip()
        enabled = {c.strip() for c in raw_ec.split(",")} if raw_ec else ALL_CHECKS
        enabled &= ALL_CHECKS  # remove unknown names

        # Model selection
        if _use_claude():
            vm, is_vis, tm = "", True, ""
            log(jid,"Claude API mode — fast and accurate",2)
        else:
            vm, is_vis = _pick_vision()
            tm = _pick_text()
            log(jid, f"Ollama | Vision:{vm} ({'capable' if is_vis else 'text-only'}) | Text:{tm} | Mode:{audit_mode.upper()}", 2)

        all_issues   = []
        shop_desc    = ""
        ref_desc     = ""
        shop_img_b64 = ""
        bbs_rows     = []

        # ── PASS 1: BBS ─────────────────────────────────────
        if bbs_path and bbs_path.exists():
            log(jid, f"Parsing BBS: {bbs_path.name}", 5)
            ext = bbs_path.suffix.lower()
            if ext in (".xlsx",".xls",".xlsm"):
                log(jid,"  Direct Excel parse (no AI needed)...")
                try:
                    bbs_rows = parse_bbs_excel(bbs_path)
                    log(jid, f"  ✓ {len(bbs_rows)} bar mark rows parsed", 15)
                    print(f"[{jid}] BBS: {len(bbs_rows)} rows parsed", flush=True)
                except Exception as e:
                    log(jid, f"  ⚠ Excel parse error: {e} — trying AI fallback")
                    bbs_rows = parse_bbs_ai(bbs_path, vm, tm, is_vis, lambda m: log(jid,m))
            else:
                bbs_rows = parse_bbs_ai(bbs_path, vm, tm, is_vis, lambda m: log(jid,m))
                log(jid, f"  ✓ {len(bbs_rows)} rows from AI parse", 15)

            # Deterministic checks (instant, no AI)
            det_issues = deterministic_bbs_checks(bbs_rows)
            all_issues.extend(det_issues)
            log(jid, f"  ✓ Deterministic checks: {len(det_issues)} issue(s) found", 18)
            try: bbs_path.unlink()
            except: pass

        if cancelled(jid): return

        # ── PASS 2: Reference drawing ─────────────────────────
        if ref_path and ref_path.exists():
            log(jid, f"Reading reference: {ref_path.name}", 20)
            ref_desc = describe_drawing(ref_path, audit_mode, vm, is_vis,
                                        REF_PROMPT, lambda m: log(jid,m))
            log(jid, f"  ✓ Reference: {len(ref_desc)} chars", 30)
            try: ref_path.unlink()
            except: pass

        if cancelled(jid): return

        # ── PASS 3: Shop drawing ──────────────────────────────
        log(jid, f"Reading shop drawing: {shop_name}", 32)
        shop_ext = shop_path.suffix.lower()
        shop_desc = describe_drawing(shop_path, audit_mode, vm, is_vis,
                                     DRAW_PROMPT, lambda m: log(jid,m))
        log(jid, f"  ✓ Shop drawing: {len(shop_desc)} chars", 50)
        print(f"[{jid}] Shop description ready: {len(shop_desc)} chars", flush=True)

        # Save first image for annotation
        try:
            if shop_ext == ".pdf":
                imgs = pdf_to_images(shop_path, max_pages=1, max_dim=1024)
                if imgs: shop_img_b64 = imgs[0]["b64"]
            else:
                img = Image.open(shop_path)
                shop_img_b64 = _compress(img, 1024)
        except: pass

        if cancelled(jid): return

        # ── PASS 4: AI Checks ─────────────────────────────────
        completed_ai = 0

        def _run_ai_check(label, prompt_text, pct_start, pct_end):
            nonlocal completed_ai
            if cancelled(jid): return
            char_count = len(prompt_text)
            log(jid, f"AI Check {completed_ai+1}/{total_ai}: {label} ({char_count} chars)...", pct_start)
            print(f"\n[{jid}] ▶ {label}  ({char_count} chars, num_predict=4000)", flush=True)

            # Heartbeat — pushes elapsed time every 15s so UI never looks frozen
            _t0 = time.time()
            _stop = threading.Event()
            def _hb():
                while not _stop.is_set():
                    _stop.wait(15)
                    if not _stop.is_set():
                        el = int(time.time()-_t0)
                        msg = f"  ⏱ {label} running... {el}s elapsed (model still thinking)"
                        log(jid, msg)
                        print(f"[{jid}]   {el}s elapsed", flush=True)
            threading.Thread(target=_hb, daemon=True).start()

            try:
                raw = ai_text(prompt_text, tm, timeout=1800, num_predict=4000)
                _stop.set()
                elapsed = int(time.time()-_t0)
                print(f"[{jid}] ✓ {label} done in {elapsed}s — {len(raw)} chars", flush=True)
                obj, err = _parse_json(raw)
                if err:
                    log(jid, f"  ⚠ Parse error: {err[:80]}")
                    print(f"[{jid}] ⚠ Parse error: {err[:120]}", flush=True)
                    return
                items = (obj.get("checks") or []) if isinstance(obj,dict) else []
                # Spread bbox for items defaulted to 0,0 — distribute across drawing
                zero_items = [x for x in items if x.get("bbox") and
                              abs(x["bbox"].get("x",0))<0.02 and abs(x["bbox"].get("y",0))<0.02]
                if zero_items:
                    # Assign positions based on check type
                    POS = {"cover_missing":(0.6,0.15),"concrete_dim_missing":(0.6,0.30),
                           "label_multiplier":(0.25,0.40),"missing_callout":(0.75,0.85),
                           "calc_error":(0.40,0.55),"dia_inconsistency":(0.30,0.20),
                           "dia_vs_rules":(0.30,0.65),"bbs_qty":(0.15,0.70),
                           "bbs_length":(0.15,0.80),"bbs_shape":(0.50,0.75),
                           "bbs_transport":(0.55,0.85),"bbs_rounding":(0.65,0.75),
                           "plan_vs_shop":(0.35,0.35),"member_mark":(0.45,0.65)}
                    used = {}
                    for it in zero_items:
                        chk = it.get("check","")
                        px, py_ = POS.get(chk,(0.5,0.5))
                        cnt = used.get(chk,0)
                        it["bbox"] = {"x":min(0.9,px+cnt*0.08),"y":min(0.9,py_+cnt*0.06),
                                      "w":0.12,"h":0.07}
                        used[chk] = cnt+1
                all_issues.extend(items)
                n_err  = sum(1 for x in items if x.get("severity")=="error")
                n_warn = sum(1 for x in items if x.get("severity")=="warning")
                log(jid, f"  ✓ {label}: {n_err} error(s), {n_warn} warning(s) in {elapsed}s", pct_end)
                print(f"[{jid}]   → {n_err} errors, {n_warn} warnings", flush=True)
            except Exception as e:
                _stop.set()
                log(jid, f"  ⚠ {label} failed: {e}")
                print(f"[{jid}] ⚠ {label} FAILED: {e}", flush=True)
            completed_ai += 1

        # ── Build modular checks based on enabled set ────────
        checks_to_run = []

        # DRAWING check — cover/dims/labels/calcs/missing
        draw_enabled = enabled & {"label_multiplier","general_arrangements","lap_qty","missing_bar"}
        if draw_enabled:
            checks_to_run.append((
                f"Drawing ({', '.join(sorted(draw_enabled))})",
                _build_drawing_prompt(shop_desc, std, rules, draw_enabled), 52, 65))

        # BBS check
        bbs_enabled = enabled & {"bbs_check","bar_length"}
        if bbs_rows and bbs_enabled:
            checks_to_run.append((
                f"BBS ({', '.join(sorted(bbs_enabled))})",
                _build_bbs_prompt(bbs_rows, shop_desc, std, bbs_enabled), 66, 78))

        # REFERENCE check
        ref_enabled = enabled & {"bar_dia","spacing"}
        if ref_desc and ref_enabled:
            checks_to_run.append((
                f"Reference ({', '.join(sorted(ref_enabled))})",
                _build_ref_prompt(ref_desc, shop_desc, std, rules, ref_enabled), 80, 92))
        elif ref_enabled and not ref_desc:
            log(jid,"  ⚠ bar_dia/spacing checks need a reference drawing — skipping",52)

        total_ai = len(checks_to_run)
        log(jid, f"Running {total_ai} AI check(s): {', '.join(c[0] for c in checks_to_run)}", 51)
        print(f"\n[{jid}] {'─'*50}", flush=True)
        print(f"[{jid}] CHECKS TO RUN: {total_ai}", flush=True)
        for ci,(lbl,_,p1,p2) in enumerate(checks_to_run,1):
            print(f"[{jid}]   {ci}. {lbl}", flush=True)
        print(f"[{jid}] DATA: shop_desc={len(shop_desc)}c | bbs_rows={len(bbs_rows)} | ref_desc={len(ref_desc)}c", flush=True)
        print(f"[{jid}] RULES: covers={rules.get('r_cover_beam','-')}/{rules.get('r_cover_slab','-')}/{rules.get('r_cover_col','-')}mm | bar_sizes={rules.get('r_bar_sizes','-')} | max_link={rules.get('r_max_link_spacing','-')}mm", flush=True)
        print(f"[{jid}] ENGINE: {_engine()} | mode={audit_mode} | std={std}", flush=True)
        print(f"[{jid}] {'─'*50}\n", flush=True)

        for label, prompt_text, p1, p2 in checks_to_run:
            _run_ai_check(label, prompt_text, p1, p2)

        if cancelled(jid): return

        # ── Annotate drawing ──────────────────────────────────
        annotated_b64 = ""
        if shop_img_b64 and all_issues:
            log(jid, "Annotating drawing with issue markers...", 93)
            try:
                annotated_b64 = annotate(shop_img_b64, all_issues)
                log(jid, f"  ✓ {len(all_issues)} markers placed", 95)
            except Exception as e:
                log(jid, f"  ⚠ Annotation: {e}")

        # ── Summary ───────────────────────────────────────────
        errors   = sum(1 for x in all_issues if x.get("severity")=="error")
        warnings = sum(1 for x in all_issues if x.get("severity")=="warning")
        grade    = "FAIL" if errors>0 else ("REVIEW" if warnings>0 else "PASS")

        result = {
            "issues":       all_issues,
            "summary":      {"critical":errors,"warnings":warnings,
                             "total":len(all_issues),"grade":grade},
            "annotated":    annotated_b64,
            "raw_drawing":  shop_img_b64,
            "shop_name":    shop_name,
            "has_ref":      bool(ref_desc),
            "has_bbs":      bool(bbs_rows),
            "bbs_rows":     bbs_rows,
            "mode":         audit_mode,
            "backend":      "claude" if _use_claude() else f"ollama:{tm}",
        }

        log(jid, f"✓ Audit complete: {errors} error(s), {warnings} warning(s) — {grade}", 100)
        print(f"\n[{jid}] ✅ COMPLETE: {errors} errors, {warnings} warnings — {grade}\n", flush=True)
        with LOCK:
            JOBS[jid]["status"] = "done"
            JOBS[jid]["result"] = result

    except Exception as e:
        import traceback; traceback.print_exc()
        push(jid,"error",str(e))
        with LOCK: JOBS[jid]["status"]="error"; JOBS[jid]["error"]=str(e)
    finally:
        try: shop_path.unlink()
        except: pass

# ═══════════════════════════════════════════════════════════
# SECTION 11 — FLASK ROUTES
# ═══════════════════════════════════════════════════════════

@app.after_request
def cors(r):
    r.headers["Access-Control-Allow-Origin"]  = "*"
    r.headers["Access-Control-Allow-Headers"] = "Content-Type"
    r.headers["Access-Control-Allow-Methods"] = "GET,POST,DELETE,OPTIONS"
    return r

@app.route("/", defaults={"p":""}, methods=["OPTIONS"])
@app.route("/<path:p>",             methods=["OPTIONS"])
def opt(p): return "",204

@app.route("/")
def index():
    if FRONTEND and FRONTEND.exists():
        return send_file(str(FRONTEND), mimetype="text/html")
    return "<h2>Frontend not found.</h2><p>Place rebar_auditor_v2.html next to app_v2.py</p>",404

@app.route("/status")
def status():
    e = _engine()
    if e == "claude":
        return jsonify({"backend":"claude","engine":"claude","ready":True,
            "vision_model":"claude-opus-4-5","text_model":"claude-sonnet-4-6","is_vision":True})
    if e == "openai":
        return jsonify({"backend":"openai","engine":"openai","ready":True,
            "vision_model":"gpt-4o","text_model":"gpt-4o","is_vision":True})
    if e == "gemini":
        return jsonify({"backend":"gemini","engine":"gemini","ready":True,
            "vision_model":"gemini-2.5-flash","text_model":"gemini-2.5-flash","is_vision":True})
    # Ollama
    ok  = _ollama_ok()
    ms  = _ollama_models() if ok else []
    vm, iv = _pick_vision() if ok else ("",False)
    tm  = _pick_text()      if ok else ""
    return jsonify({"backend":"ollama","engine":"ollama","ollama_running":ok,"models":ms,
                    "vision_model":vm,"is_vision":iv,"text_model":tm,
                    "ready": ok and bool(ms),
                    "api_available": bool(_CLAUDE_KEY or _OPENAI_KEY or _GEMINI_KEY),
                    "suggested_engine": ("claude" if _CLAUDE_KEY else
                                         "openai" if _OPENAI_KEY else
                                         "gemini" if _GEMINI_KEY else "")})

@app.route("/audit/start", methods=["POST"])
def audit_start():
    if "drawing" not in request.files: return jsonify({"error":"No drawing uploaded"}),400

    # Check engine
    if _use_claude():
        pass  # always ready
    elif not _ollama_ok():
        return jsonify({"error":"Ollama not running. Run: ollama serve"}),503
    elif not _ollama_models():
        return jsonify({"error":"No Ollama models. Run: ollama pull llava:34b"}),503

    df  = request.files["drawing"]
    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    jid = uuid.uuid4().hex[:8]
    shop_path = UPLOAD_DIR/f"shop_{ts}_{jid}{Path(df.filename).suffix.lower()}"
    df.save(str(shop_path))

    ref_path = None
    rf = request.files.get("reference")
    if rf and rf.filename:
        ref_path = UPLOAD_DIR/f"ref_{ts}_{jid}{Path(rf.filename).suffix.lower()}"
        rf.save(str(ref_path))

    bbs_path = None
    bf = request.files.get("bbs")
    if bf and bf.filename:
        bbs_path = UPLOAD_DIR/f"bbs_{ts}_{jid}{Path(bf.filename).suffix.lower()}"
        bf.save(str(bbs_path))

    cond = {k:request.form.get(k,"") for k in
            ("standard","audit_mode","custom_notes","enabled_checks",
             "r_cover_beam","r_cover_slab","r_cover_col",
             "r_bar_sizes","r_max_link_spacing")}
    cond["standard"]   = cond["standard"]   or "BS8666"
    cond["audit_mode"] = cond["audit_mode"] or "smart"

    with LOCK: JOBS[jid] = {"status":"running","events":[],"result":None,"cancelled":False}
    threading.Thread(target=run_job,
        args=(jid, shop_path, ref_path, bbs_path, cond, df.filename), daemon=True).start()
    return jsonify({"job_id":jid})

@app.route("/audit/stream/<jid>")
def audit_stream(jid):
    def gen():
        import time; cursor=0
        while True:
            with LOCK:
                job=JOBS.get(jid)
                if not job: yield f"data:{json.dumps({'type':'error','data':'Not found'})}\n\n"; return
                evs=job["events"][cursor:]; cursor+=len(evs); st=job["status"]
            for ev in evs: yield f"data:{json.dumps(ev)}\n\n"
            if st in ("done","error"): yield f"data:{json.dumps({'type':'done','data':st})}\n\n"; return
            time.sleep(0.4)
    return Response(stream_with_context(gen()), mimetype="text/event-stream",
        headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})

@app.route("/audit/result/<jid>")
def audit_result(jid):
    with LOCK: job=JOBS.get(jid)
    if not job:                  return jsonify({"error":"Not found"}),404
    if job["status"]=="running": return jsonify({"status":"running"}),202
    if job["status"]=="error":   return jsonify({"error":job.get("error","Unknown")}),500
    return jsonify({"success":True,"results":job["result"]})


@app.route("/set-engine", methods=["POST"])
def set_engine_route():
    global _ENGINE, _CLAUDE_KEY, _OPENAI_KEY, _GEMINI_KEY
    data   = request.json or {}
    engine = data.get("engine","ollama").lower()
    key    = data.get("key","").strip()

    if engine == "claude":
        if not key: return jsonify({"error":"Claude API key required (sk-ant-...)"}),400
        _CLAUDE_KEY = key; _ENGINE = "claude"
        return jsonify({"ok":True,"engine":"claude","msg":"Claude API active (claude-opus-4-5 vision, claude-sonnet-4-6 text)"})

    elif engine == "openai":
        if not key: return jsonify({"error":"OpenAI API key required (sk-...)"}),400
        _OPENAI_KEY = key; _ENGINE = "openai"
        return jsonify({"ok":True,"engine":"openai","msg":"OpenAI active (gpt-4o vision + text)"})

    elif engine == "gemini":
        if not key: return jsonify({"error":"Google API key required"}),400
        _GEMINI_KEY = key; _ENGINE = "gemini"
        return jsonify({"ok":True,"engine":"gemini","msg":"Gemini 2.5 Flash active (vision + text)"})

    else:  # ollama
        _ENGINE = "ollama"
        return jsonify({"ok":True,"engine":"ollama","msg":"Switched to Ollama (local)"})

# Legacy aliases
@app.route("/set-api-key",  methods=["POST"])
def set_api_key():
    return set_engine_route()
@app.route("/clear-api-key", methods=["POST"])
def clear_api_key():
    global _ENGINE; _ENGINE="ollama"; return jsonify({"ok":True,"msg":"Ollama mode"})

@app.route("/audit/cancel/<jid>", methods=["POST"])
def audit_cancel(jid):
    with LOCK:
        job=JOBS.get(jid)
        if job and job["status"]=="running":
            job["cancelled"]=True; job["status"]="error"; job["error"]="Cancelled by user"
            return jsonify({"ok":True})
    return jsonify({"error":"Not found"}),404

@app.route("/audit/localize/<jid>", methods=["POST"])
def audit_localize(jid):
    """Re-locate zero-position issues using vision model."""
    with LOCK: job=JOBS.get(jid)
    if not job or job["status"]!="done": return jsonify({"error":"Job not ready"}),404
    result = job.get("result",{})
    draw_b64 = result.get("raw_drawing","")
    if not draw_b64: return jsonify({"error":"No drawing image stored"}),404

    vm, is_vis = _pick_vision() if not _use_claude() else ("",True)
    zero = [(i,x) for i,x in enumerate(result.get("issues",[]))
            if x.get("bbox") and abs(x["bbox"].get("x",0))<0.02 and abs(x["bbox"].get("y",0))<0.02]
    if not zero: return jsonify({"locs":[]})

    BATCH=15; updated=[]
    for bs in range(0, min(len(zero),45), BATCH):
        batch = zero[bs:bs+BATCH]
        lines = "\n".join(f"{i+1}. [{it.get('element_type','?')} {it.get('element_id','')}] "
                          f"{(it.get('issue',''))[:80]}" for i,(_,it) in enumerate(batch))
        prompt = (f"Rebar shop drawing. Locate {len(batch)} issues precisely.\n\nISSUES:\n{lines}\n\n"
                  'Return JSON: {"locs":[{"n":1,"bbox":{"x":0.0,"y":0.0,"w":0.1,"h":0.08}}]}\n'
                  "x,y=top-left 0-1, w,h=size. Different position per issue. Never x=0,y=0 unless truly top-left.")
        try:
            raw = ai_vision(draw_b64, prompt, vm, is_vis, timeout=600, num_predict=2000)
            obj,_ = _parse_json(raw)
            if obj:
                for loc in obj.get("locs",[]):
                    gi = bs + loc.get("n",0) - 1
                    if 0<=gi<len(zero) and loc.get("bbox"):
                        bb=loc["bbox"]
                        if not (abs(bb.get("x",0))<0.02 and abs(bb.get("y",0))<0.02):
                            real_idx = zero[gi][0]
                            result["issues"][real_idx]["bbox"] = bb
                            updated.append({"idx":real_idx,"bbox":bb})
        except: pass

    # Re-annotate
    annotated=""
    if updated and draw_b64:
        try: annotated = annotate(draw_b64, result["issues"])
        except: pass
    if annotated: result["annotated"] = annotated

    return jsonify({"locs":updated,"annotated_image":annotated})

# ═══════════════════════════════════════════════════════════
# SECTION 12 — MAIN
# ═══════════════════════════════════════════════════════════

if __name__=="__main__":
    print("="*60)
    print("  Rebar QA/QC Auditor v2.0  —  Clean Rewrite")
    print("="*60)
    e = _engine()
    if e == "claude":
        print(f"\n  ✓ Claude API  | Vision: claude-opus-4-5 | Text: claude-sonnet-4-6")
    elif e == "openai":
        print(f"\n  ✓ OpenAI API  | Vision: gpt-4o | Text: gpt-4o")
    elif e == "gemini":
        print(f"\n  ✓ Gemini 2.5  | Vision: gemini-2.5-flash | Text: gemini-2.5-flash")
    if e == "claude" or e == "openai" or e == "gemini":
        pass  # cloud engine — skip Ollama check below
    else:
        ok = _ollama_ok()
        ms = _ollama_models() if ok else []
        print(f"\n  Ollama: {'✓ running' if ok else '✗ NOT running — run: ollama serve'}")
        if ms:
            vm,iv = _pick_vision()
            tm    = _pick_text()
            print(f"  Vision: {vm} ({'✓ vision capable' if iv else '✗ text-only'})")
            print(f"  Text:   {tm}")
        else:
            print("  No models found. Run:")
            print("    ollama pull llava:34b")
            print("    ollama pull llama3.3:70b")
    if FRONTEND:
        print(f"\n  ✓ Frontend: {FRONTEND.name}")
    else:
        print("\n  ⚠ Frontend not found — place rebar_auditor_v2.html here")
    print(f"\n  ▶ Open: http://localhost:5000")
    print("="*60+"\n")
    import webbrowser, threading

    if _engine() == "ollama" and not _ollama_ok():
        print("\n  ⚠  OLLAMA NOT RUNNING — open the app and go to:")
        print("  Audit Rules > AI Engine > enter OpenAI / Gemini / Claude key")
        print("  A quick key entry box also appears on the main upload page.\n")
    def _open():
        import time; time.sleep(2.5)
        webbrowser.open("http://localhost:5000")
    threading.Thread(target=_open, daemon=True).start()
    print("\n  Browser will open automatically at http://localhost:5000\n", flush=True)
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False, threaded=True)
